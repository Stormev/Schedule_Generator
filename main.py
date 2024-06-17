import sys
import os
import traceback
from openpyxl.styles.borders import Side
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import json
import PyQt5.QtWidgets as qwidget
from PyQt5.QtGui import QFont
from PyQt5 import uic

# При добавлении нового праметра, удаляем старый json файл
default_settings = {'json_encode': True,
                    'rows_count': 20,
                    'column_count': 20,
                    'program_font': 'Segoe UI',
                    'negative_schedule': False,
                    'saturday_enabled': True,
                    'max_hours_week': 36
                    }


def settings_create_if_notExist():
    if not os.path.isfile('data/settings.json'):
        with open('data/settings.json', 'w', encoding='utf-8') as f:
            f.write(json.dumps(default_settings, indent=4, ensure_ascii=False))


# Проверка на существование файла с настройками
settings_create_if_notExist()


def get_settings():
    cur_path = 'data/settings.json'
    if os.path.exists(cur_path):
        with open(cur_path, 'r', encoding='utf-8') as f:
            return json.loads(f.read())
    settings_create_if_notExist()
    return default_settings


def settings_reset():
    cur_path = 'data/settings.json'
    if os.path.exists(cur_path):
        with open(cur_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(default_settings, indent=4, ensure_ascii=False))


# Виджет настройки
class UiSettigns(qwidget.QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('UI/settings.ui', self)
        self.setWindowTitle('Настройки')
        self.setMaximumSize(457, 300)
        # конфиг
        self.current_settings = get_settings()
        self.spinBox_row.setValue(self.current_settings['rows_count'])
        self.spinBox_col.setValue(self.current_settings['column_count'])
        self.fontComboBox.setCurrentFont(QFont(self.current_settings['program_font']))
        self.checkBox_negative.setChecked(self.current_settings['negative_schedule'])
        self.checkBox_saturday.setChecked(self.current_settings['saturday_enabled'])

        self.hide_all()
        # Выбор раздела
        self.listWidget.itemClicked.connect(self.chose_frame)

        # Взаимодействие
        self.spinBox_row.valueChanged.connect(lambda x: self.change_settings('rows_count', self.spinBox_row.value()))
        self.spinBox_col.valueChanged.connect(lambda x: self.change_settings('column_count', self.spinBox_col.value()))

        self.fontComboBox.currentFontChanged.connect(
            lambda x: self.change_settings('program_font', self.fontComboBox.currentFont().family()))

        self.checkBox_negative.stateChanged.connect(
            lambda x: self.change_settings('negative_schedule', self.checkBox_negative.isChecked()))
        self.checkBox_saturday.stateChanged.connect(
            lambda x: self.change_settings('saturday_enabled', self.checkBox_saturday.isChecked()))

        self.init()

    def init(self):
        self.show()

    def hide_all(self):
        self.frame_row_col.hide()
        self.frame_font.hide()
        self.frame_add.hide()

    def change_settings(self, key, value):
        self.current_settings = get_settings()
        self.current_settings[key] = value

        cur_path = 'data/settings.json'
        if os.path.exists(cur_path):
            with open(cur_path, 'w', encoding='utf-8') as file:
                file.write(json.dumps(self.current_settings, indent=4, ensure_ascii=False))

    def chose_frame(self):
        action = {'0': self.frame_row_col.show, '1': self.frame_font.show, '2': self.frame_add.show}
        self.hide_all()
        action.get(str(self.listWidget.currentRow()))()


# Виджет о программе
class UiAbout(qwidget.QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('UI/about.ui', self)
        self.setWindowTitle('О Программе')
        self.setMaximumSize(396, 580)
        self.init()

    def init(self):
        if os.path.exists('README.txt'):
            with open('README.txt', 'r', encoding='utf-8') as f:
                text = f.read().split('#-')
                self.textBrowser_about.setText(text[1])
                self.textBrowser_direction.setText(text[0])
        else:
            self.textBrowser_about.setText('отсутсвует README файл')
            self.textBrowser_direction.setText('отсутсвует README файл')
        self.show()


# Главное окно
class FrameMain(qwidget.QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('UI/main.ui', self)
        self.setWindowTitle('Генератор Расписания')
        # self.setMaximumSize(1200, 700)

        # Сохранение и загрузка файла происходит в порядке расстановки в списке,
        # НЕ МЕНЯТЬ ПОРЯДОК!
        self.itemGroups = [
            # постоянный набор
            [self.table_add, self.table_rooms, self.table_teachers],
            # Груповой набор
            [self.table_groups, self.table_hours, self.table_addHours],
            # Привязка
            [self.table_binding],
        ]

        self.itemWeek = [self.table_main_1, self.table_main_2, self.table_main_3, self.table_main_4, self.table_main_5,
                         self.table_main_6
                         ]

        self.frame = None
        self.initUI()

    def initUI(self):
        # начальный конфиг
        self.listWidget_file.hide()
        # Меню бар кнопки
        self.btn_file.clicked.connect(self.show_file)
        self.btn_settings.clicked.connect(self.ui_settings)
        self.btn_generate.clicked.connect(self.generate)  # -----
        self.btn_about.clicked.connect(self.ui_about)
        self.listWidget_file.itemClicked.connect(lambda x: self.file_edit_widget(x.text()))

        # update ивент
        self.table_teachers.itemChanged.connect(self.update)
        self.table_groups.itemChanged.connect(self.update)
        self.table_add.itemChanged.connect(self.update)

        self.show()

        # Настройка таблиц
        for group in self.itemGroups:
            for table in group:
                table.setRowCount(int(get_settings()['rows_count']))

    # Работа с файлом
    def show_file(self):
        object = self.listWidget_file
        if not object.isVisible():
            object.show()
        else:
            object.hide()

    def file_edit_widget(self, mode):

        result = {'Создать новый файл': self.newFile, 'Загрузить модель файла': self.loadFile,
                  'Сохранить модель файла': self.saveFile,
                  'Экспортировать файл в .xlsx (Exel)': self.export_to_exel()}.get(mode)
        if result:
            result()
        self.show_file()

    # Открыть настройки
    def ui_settings(self):
        self.frame = UiSettigns()
        self.frame.show()

    def create_critical_msg(self, text):
        error_dialog = qwidget.QMessageBox()
        error_dialog.setIcon(qwidget.QMessageBox.Critical)
        error_dialog.setText(text)
        error_dialog.setStandardButtons(qwidget.QMessageBox.Ok)
        error_dialog.setWindowTitle('Ошибка')
        error_dialog.exec_()

    # Очиста правых таблиц
    def clear_week(self):
        for table in self.itemWeek:
            for column in range(table.columnCount()):
                table.setHorizontalHeaderItem(column, qwidget.QTableWidgetItem(''))
            row = table.rowCount()
            col = table.columnCount()
            table.setRowCount(0)
            table.setColumnCount(0)
            table.setRowCount(row)
            table.setColumnCount(col)

    # Главная функция
    def generate(self):
        # Очистка таблицы
        self.clear_week()
        print('Запуск генерации')

        # Получаем словарь: {группа: {предмет:часов}}

        def toNewDict(old_dict: dict) -> dict:
            result = {}
            flag = False
            for key, item in old_dict.items():
                if key and item and not flag:
                    for value in item:
                        if value:
                            result[value] = {}
                    flag = True
                else:
                    result_keys = result.copy().keys()
                    for new_key, index in zip(result_keys, range(len(result_keys))):
                        if not new_key or len(item) <= index or not key:
                            break
                        result[new_key].update([(key, item[index])])
            return result

        # Требуется для инкапсуляции данных
        def get_add_hourse():
            data_add_hours = toNewDict(self.getTable(self.table_addHours))
            data_add_info = toNewDict(self.getTable(self.table_add))
            return [data_add_hours, data_add_info]

        generate_data = {'HOURS': toNewDict(self.getTable(self.table_hours)),
                         'ADD_HOURSE': get_add_hourse(),
                         'TEACHERS': toNewDict(self.getTable(self.table_teachers)),
                         'FREE_ROOMS': self.getTable(self.table_rooms),
                         'GROUPS': toNewDict(self.getTable(self.table_groups)),
                         'ATTACHED_TEACHERS': toNewDict(self.getTable(self.table_binding)),
                         'MAX_WEEK_LESSONS': {},
                         }

        def add_dict_hours():
            max_lessons = int(get_settings()['max_hours_week']) // 2
            [generate_data['MAX_WEEK_LESSONS'].update({key: max_lessons}) for key in generate_data['GROUPS'].keys()]

        add_dict_hours()
        # Если суббота включена то 6-(1-1)=6 если нет то 6-(1-0)=5
        saturday_enabled = 1 - int(get_settings()['saturday_enabled'])
        max_days = len(self.itemWeek) - saturday_enabled

        # Выбирает пару с наибольшем количеством часов
        def choise_lesson(dict_hours: dict, used_list=[]):
            temp = dict_hours.copy()
            if used_list:
                [temp.pop(item) if temp.get(item) else False for item in used_list]
                if not temp:
                    return ''
            return max(temp, key=lambda x: int(temp.get(x)))

        # ----------ГЛАВНЫЙ ЦИКЛ ГЕНЕРАЦИИ----------------------------------------------------
        # Цикл по дням недели
        for day_index in range(int(max_days)):
            day_table = self.itemWeek[day_index]
            print(f'Процесс Генерации: {day_index + 1}/{max_days + day_index}')

            # Учителя которые используются на одинаковом ряду пар
            used_teachers = {}
            # Модификатор ключа словаря сверху
            inter_mod = lambda x: 'numb_' + str(x)
            # Цикл групп
            for key, col in zip(generate_data['GROUPS'].keys(), range(len(generate_data['GROUPS']))):
                day_table.setHorizontalHeaderItem(col, qwidget.QTableWidgetItem(key))
                # Количество пар на день
                max_lessons_day = round(generate_data['MAX_WEEK_LESSONS'][key] / max_days)
                # Вычитаем часы из общей массы группы
                generate_data['MAX_WEEK_LESSONS'][key] -= max_lessons_day
                # Убераем повторения пар
                used_lessons = []
                # Список уже используемых свободных кабинетов
                used_classes = []
                # Цикл пар
                for row in range(max_lessons_day):
                    flag_priority = False
                    current_lesson, teacher, teacher_room = ''

                    # ----Проверяем дополнительные часы--
                    for extra_lesson, extra_lesson_value in generate_data['ADD_HOURSE'][1].items():
                        if int(extra_lesson_value['№Пары']) == row \
                                and int(generate_data['ADD_HOURSE'][0][key][extra_lesson]) >= 1 \
                                and int(extra_lesson_value['День Недели(1-6)']) == day_index + 1:
                            # Запись результата
                            day_table.setItem(row, col, qwidget.QTableWidgetItem(extra_lesson))
                            flag_priority = True
                            # Добавление 0-ого ряда
                            for column_table in range(day_table.columnCount()):
                                day_table.setVerticalHeaderItem(column_table,
                                                                qwidget.QTableWidgetItem(str(column_table)))
                            break
                    if flag_priority:
                        continue

                    # Выбирает пару если нет приоритетной
                    current_lesson = choise_lesson(generate_data['HOURS'][key], used_list=used_lessons)
                    cell_value = int(generate_data['HOURS'][key][current_lesson])
                    # Если количество неповторяемых пар меньше 0 то будут повторяемые
                    if cell_value <= 0:
                        used_lessons.clear()
                        current_lesson = choise_lesson(generate_data['HOURS'][key])
                        if int(generate_data['HOURS'][key][current_lesson]) <= 0:
                            current_lesson = ''
                    elif current_lesson:
                        # Получаем привязанного преподователя по предмету
                        teacher = generate_data['ATTACHED_TEACHERS'][key][current_lesson]
                        # Если преподователь уже занят
                        if used_teachers.get(inter_mod(row)) and teacher in used_teachers.get(inter_mod(row)):
                            # Цикл поиска свободного преподователя на определённой паре
                            count_stop = 10
                            while teacher in used_teachers.get(inter_mod(row)):
                                if count_stop <= 10:
                                    if row == 0 or row + 1 >= max_lessons_day:
                                        current_lesson = ''
                                    break
                                current_lesson = choise_lesson(generate_data['HOURS'][key],
                                                               used_list=used_lessons + [current_lesson])
                                teacher = generate_data['ATTACHED_TEACHERS'][key].get(current_lesson)
                    # Проверка на пустую пару
                    if current_lesson:
                        # Фиксируем часы
                        used_lessons.append(current_lesson)
                        generate_data['HOURS'][key][current_lesson] = cell_value - 2
                        # Получаем кабинет
                        teacher_room = generate_data['TEACHERS'][teacher]['Кабинет']
                        if not teacher_room:
                            # Список свободных комнат
                            rooms = self.getTable(self.table_rooms).popitem()[1]
                            # Фильтр
                            rooms = [used_room for used_room in rooms if used_room]
                            [rooms.remove(used_room) for used_room in used_classes if used_room]
                            # Выбор комнаты
                            teacher_room = rooms.pop(0)
                            used_classes.append(teacher_room)
                            # Работа со словарём used_teachers. Нужно для создания словаря {№Пары:[Преподователи]}
                        if not used_teachers.get(inter_mod(row)):
                            used_teachers[inter_mod(row)] = [teacher]
                        else:
                            used_teachers[inter_mod(row)] = used_teachers[inter_mod(row)] + [teacher]

                        # Итог дополняем информацией
                        current_lesson = f'{current_lesson}({teacher.split()[0]}) {teacher_room} каб'
                        # Запись результата
                        day_table.setItem(row, col, qwidget.QTableWidgetItem(current_lesson))
            # Максимальные дни уменьшаем дабы правильно рассчитывать количество пар на день
            max_days -= 1
        print('Генерация завершена')

    # Открыть о программе
    def ui_about(self):
        self.frame = UiAbout()
        self.frame.show()

    # Функция, которая возвращает словарь {название колонки: [массив зачений]}
    def getTable(self, object, isdict=False):
        packet = {}
        # Возвращает список
        if not isdict:
            if object.horizontalHeaderItem(0):
                for header in range(object.horizontalHeader().count()):
                    if not object.horizontalHeaderItem(header):
                        break
                    packet[object.horizontalHeaderItem(header).text()] = []
            else:
                for val in range(object.horizontalHeader().count()):
                    packet[f'{val}'] = []

            for directory, column in zip(packet.keys(), range(object.columnCount())):
                for row in range(0, object.rowCount()):
                    if directory and object.item(row, column):
                        packet.get(directory).append(object.item(row, column).text())
        # Возвращает словарь, где номер_ряда=значение
        else:
            if object.horizontalHeaderItem(0):
                for header in range(object.horizontalHeader().count()):
                    if not object.horizontalHeaderItem(header):
                        break
                    packet[object.horizontalHeaderItem(header).text()] = {}
            else:
                for val in range(object.horizontalHeader().count()):
                    packet[f'{val}'] = {}

            for directory, column in zip(packet.keys(), range(object.columnCount())):
                for row in range(0, object.rowCount()):
                    if directory and object.item(row, column):
                        packet.get(directory)[f'{row}'] = object.item(row, column).text()
        return packet

    # Очищает всё
    def newFile(self):
        for index in range(len(self.itemGroups)):
            for table in self.itemGroups[index]:
                for column in range(table.columnCount()):
                    for row in range(table.rowCount()):
                        table.setItem(row, column, qwidget.QTableWidgetItem(''))

    def saveFile(self):
        # Сохранение модели в json
        primary_data = {}
        pathName, type = qwidget.QFileDialog.getSaveFileName(self, directory="C://", filter='JSON (*.json)')
        if pathName:
            for group, number in zip(self.itemGroups, range(len(self.itemGroups))):
                packet_name = f'Packet_{number}'
                primary_data[packet_name] = []
                for item in group:
                    primary_data.get(packet_name).append(self.getTable(item))
        if primary_data:
            with open(pathName, 'w') as f:
                f.write(json.dumps(primary_data, ensure_ascii=get_settings()['json_encode'], indent=4))

    def loadFile(self):
        current_file, file_type = qwidget.QFileDialog.getOpenFileName(self, directory='C://', filter='JSON (*.json)')
        self.newFile()
        # Проверка на существование файла
        if current_file and os.path.exists(current_file):
            with open(current_file, 'r', encoding='utf-8') as f:
                current_data = json.loads(f.read())
                if current_data:
                    # Проходит по пакетам (административный, групповой)
                    for item, index in zip(current_data.values(), range(len(self.itemGroups))):
                        # Проходит по существующим таблицам
                        for table, tableIndex in zip(self.itemGroups[index], range(len(self.itemGroups[index]))):
                            # self.update()
                            # Проходит по колонкам таблицы
                            self.update()
                            for column in range(table.columnCount()):
                                # Получаем из данных, где ключ - название колонки
                                value = item[tableIndex].get(table.horizontalHeaderItem(column).text())
                                if value:
                                    for row in range(0, len(value)):
                                        table.setItem(row, column, qwidget.QTableWidgetItem(f'{value[row]}'))
        self.update()

    # конвертация в .xlsx
    def export_to_exel(self):
        current_data = []

        # Дни недели
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

        # Стиль
        cell_border = Side(border_style="thin", color="000000")
        cell_head_color = PatternFill(start_color='879bf5', end_color='879bf5', fill_type="solid")
        cell_corner_color = PatternFill(start_color='a7a7a8', end_color='a7a7a8', fill_type="solid")
        cell_height, cell_width = 25, 15

        # Получаем значения
        for table in self.itemWeek:
            current_data.append(self.getTable(table, isdict=True))

        if current_data:
            work_b = Workbook()
            ws = work_b.active
            # Цикл дней
            for day, number in zip(current_data, range(len(current_data))):
                ws.title = days[number]
                ws.cell(row=1, column=1).fill = cell_corner_color
                ws.column_dimensions['A'].width = cell_width
                ws['A1'] = ' Пара -- Группа '
                col = 2
                # групп-пар
                for group, item in day.items():
                    if item:
                        ws.cell(row=1, column=col).fill = cell_head_color
                        ws.cell(row=1, column=col).value = group

                        for row, val in item.items():
                            ws.cell(row=int(row) + 2, column=col).value = val
                            ws.cell(row=int(row) + 2, column=1).value = int(row)
                    col += 1
                if number < 5:
                    ws = work_b.create_sheet()

            work_b.save(filename='test.xlsx')

    # Вызывается при обновлении данных в таблицах
    def update(self):

        # !! Инкапсуляция неполучится, частный случай

        # Добавляет специализацию к таблице часам для групп

        # спец функция для загрузки существующей информации (для сохранения данных в таблице при обновлении)
        def load_table(table, data: dict):
            for col in range(table.horizontalHeader().count()):
                for row in range(0, table.rowCount()):
                    if not table.horizontalHeaderItem(col):
                        break
                    value = data.get(table.horizontalHeaderItem(col).text())
                    if not value or row >= len(value):
                        break
                    value = value[row]
                    table.setItem(row, col, qwidget.QTableWidgetItem(value))

        # Обновление основных часов
        def load_table_hours():
            # Специализации
            already = []
            table_data = self.getTable(self.table_hours)

            for row in range(0, self.table_teachers.rowCount()):
                value = self.table_teachers.item(row, 1)
                if value and value.text():
                    value = value.text().strip().lower()
                    if value and value not in already:
                        already.append(value)
                        self.table_hours.setColumnCount(len(already) + 1)
                        self.table_hours.setHorizontalHeaderItem(row + 1, qwidget.QTableWidgetItem(value))

            # Группы
            already = []
            for row in range(0, self.table_groups.rowCount()):
                value = self.table_groups.item(row, 0)
                if value and value.text() not in already and value.text():
                    value = value.text()
                    already.append(value)
                    self.table_hours.setRowCount(len(already))
                    self.table_hours.setItem(row, 0, qwidget.QTableWidgetItem(value))

            load_table(table=self.table_hours, data=table_data)

        # Обновление доп часов
        def load_table_addhours():
            # Группа
            already = []
            table_data = self.getTable(self.table_addHours)

            for row in range(0, self.table_groups.rowCount()):
                value = self.table_groups.item(row, 0)
                if value and value.text() not in already and value.text():
                    value = value.text()
                    already.append(value)
                    self.table_addHours.setRowCount(len(already) + 1)
                    self.table_addHours.setItem(row, 0, qwidget.QTableWidgetItem(value))
            # Доп
            already = []
            for row in range(0, self.table_add.rowCount()):
                value = self.table_add.item(row, 0)
                if value and value.text() not in already:
                    value = value.text()
                    already.append(value)
                    self.table_addHours.setColumnCount(len(already) + 1)
                    self.table_addHours.setHorizontalHeaderItem(row + 1, qwidget.QTableWidgetItem(value))

            load_table(table=self.table_addHours, data=table_data)

        # Обновление привязок
        def load_table_binding():
            already = []
            table_data = self.getTable(self.table_binding)
            for row in range(0, self.table_teachers.rowCount()):
                value = self.table_teachers.item(row, 1)
                if value and value.text():
                    value = value.text().strip().lower()
                    if value and value not in already:
                        already.append(value)
                        self.table_binding.setColumnCount(len(already) + 1)
                        self.table_binding.setHorizontalHeaderItem(row + 1, qwidget.QTableWidgetItem(value))

            # Группы
            already = []
            for row in range(0, self.table_groups.rowCount()):
                value = self.table_groups.item(row, 0)
                if value and value.text() not in already and value.text():
                    value = value.text()
                    already.append(value)
                    self.table_binding.setRowCount(len(already))
                    self.table_binding.setItem(row, 0, qwidget.QTableWidgetItem(value))

                    load_table(table=self.table_binding, data=table_data)

        load_table_addhours()
        load_table_hours()
        load_table_binding()
        try:
            pass
        except Exception as err:
            self.create_critical_msg(f'Ошибка обновления таблиц: {err}')


def excepthook(exc_type, exc_value, exc_tb):
    tb = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print("Oбнаружена ошибка !:", tb)


sys.excepthook = excepthook

# app -приложение frame_main - главное окно
app = None
frame_main = None


# билд инициализации
def __initFrame__():
    global app
    global frame_main

    app = qwidget.QApplication(sys.argv)
    frame_main = FrameMain()
    sys.exit(app.exec_())


if __name__ == '__main__':
    __initFrame__()
    input()
